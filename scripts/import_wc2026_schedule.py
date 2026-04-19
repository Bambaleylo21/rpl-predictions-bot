from __future__ import annotations

import argparse
import asyncio
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models import Match, Tournament

WC_CODE = "WC2026"
WC_NAME = "ЧМ 2026"


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip().lower().replace("ё", "е"))


def _round_to_number(raw: Any) -> int | None:
    txt = _norm(raw)
    if not txt:
        return None
    if txt.isdigit():
        num = int(txt)
        if 1 <= num <= 9:
            return num
    if "тур 1" in txt:
        return 1
    if "тур 2" in txt:
        return 2
    if "тур 3" in txt:
        return 3
    if "1/16" in txt:
        return 4
    if "1/8" in txt:
        return 5
    if "четверть" in txt:
        return 6
    if "полуфин" in txt:
        return 7
    if "3" in txt and "мест" in txt:
        return 8
    if "финал" in txt:
        return 9
    return None


def _parse_match_pair(raw: Any) -> tuple[str, str] | None:
    txt = str(raw or "").strip()
    if not txt:
        return None
    # ВАЖНО: не разделяем по обычному дефису внутри названий команд
    # (например, "Кот-д’Ивуар", "Кабо-Верде").
    # Ищем только реальные разделители между командами.
    delimiters = [" — ", " – ", " - ", "—", "–"]
    for delim in delimiters:
        if delim in txt:
            parts = [x.strip() for x in txt.split(delim, 1)]
            if len(parts) == 2 and parts[0] and parts[1]:
                return parts[0], parts[1]
    return None


def _parse_date(raw: Any) -> datetime.date | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        try:
            dt = from_excel(raw)
            if isinstance(dt, datetime):
                return dt.date()
            if isinstance(dt, date):
                return dt
        except Exception:
            pass
    if isinstance(raw, datetime):
        return raw.date()
    txt = str(raw).strip()
    if not txt:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            pass
    return None


def _parse_time(raw: Any) -> time | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        try:
            dt = from_excel(raw)
            if isinstance(dt, datetime):
                return dt.time().replace(second=0, microsecond=0)
            if isinstance(dt, time):
                return dt.replace(second=0, microsecond=0)
        except Exception:
            pass
    if isinstance(raw, datetime):
        return raw.time().replace(second=0, microsecond=0)
    if isinstance(raw, time):
        return raw.replace(second=0, microsecond=0)
    txt = str(raw).strip()
    if not txt:
        return None
    for fmt in ("%H:%M", "%H.%M"):
        try:
            return datetime.strptime(txt, fmt).time()
        except ValueError:
            pass
    return None


def _extract_kickoff_from_row(row: list[Any]) -> datetime | None:
    date_part = None
    time_part = None
    for cell in row:
        if isinstance(cell, datetime):
            return cell.replace(second=0, microsecond=0)
        if date_part is None:
            date_part = _parse_date(cell)
        if time_part is None:
            time_part = _parse_time(cell)
    if date_part is None or time_part is None:
        return None
    return datetime.combine(date_part, time_part)


def _detect_group_from_row(row: list[Any], skip_indexes: set[int]) -> str | None:
    for idx, cell in enumerate(row):
        if idx in skip_indexes:
            continue
        txt = str(cell or "").strip()
        if not txt:
            continue
        n = _norm(txt)
        if n.startswith("группа ") or n.startswith("group "):
            return txt
        # Частый формат просто "A", "B", ..., "H"
        if len(txt) == 1 and txt.upper() in {"A", "B", "C", "D", "E", "F", "G", "H"}:
            return txt.upper()
    return None


def _detect_columns(header_cells: list[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, cell in enumerate(header_cells):
        h = _norm(cell)
        if not h:
            continue
        if "груп" in h:
            out["group"] = i
        elif "раунд" in h or "тур" in h or "этап" in h or "round" in h or "stage" in h:
            out["round"] = i
        elif "дат" in h or "date" in h or "day" in h:
            out["date"] = i
        elif "врем" in h or "начал" in h or "time" in h or "kickoff" in h:
            out["time"] = i
        elif "матч" in h or "match" in h:
            out["match"] = i
        elif "дом" in h or "home" in h:
            out["home"] = i
        elif "гост" in h or "away" in h:
            out["away"] = i
        elif h == "команды":
            out["match"] = i
    return out


async def run_import(xlsx_path: Path, clear_wc_matches: bool = False) -> None:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)

    rows: list[tuple[list[Any], dict[str, int], str, int]] = []
    for ws in wb.worksheets:
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows:
            continue
        header_idx = None
        mapping: dict[str, int] = {}
        for i, r in enumerate(sheet_rows[:100]):
            m = _detect_columns(list(r))
            has_round = "round" in m
            has_date = "date" in m
            has_time = "time" in m
            has_match = "match" in m or ("home" in m and "away" in m)
            if has_round and has_date and has_time and has_match:
                header_idx = i
                mapping = m
                break
        if header_idx is None:
            continue
        rows.append((sheet_rows, mapping, ws.title, header_idx))

    parsed: list[dict[str, Any]] = []
    skipped_no_match = 0
    skipped_bad_round = 0
    skipped_bad_datetime = 0

    for sheet_rows, mapping, _sheet_title, header_idx in rows:
        for r in sheet_rows[header_idx + 1 :]:
            if r is None:
                continue
            round_raw = r[mapping["round"]] if mapping.get("round") is not None else None
            round_number = _round_to_number(round_raw)
            if round_number is None:
                if any(x is not None and str(x).strip() for x in r):
                    skipped_bad_round += 1
                continue

            pair = None
            if mapping.get("match") is not None:
                match_raw = r[mapping["match"]]
                pair = _parse_match_pair(match_raw)
            elif mapping.get("home") is not None and mapping.get("away") is not None:
                home_raw = str(r[mapping["home"]] or "").strip()
                away_raw = str(r[mapping["away"]] or "").strip()
                if home_raw and away_raw:
                    pair = (home_raw, away_raw)
            if pair is None:
                skipped_no_match += 1
                continue

            date_raw = r[mapping["date"]] if mapping.get("date") is not None else None
            time_raw = r[mapping["time"]] if mapping.get("time") is not None else None
            d = _parse_date(date_raw)
            t = _parse_time(time_raw)
            if d is None or t is None:
                skipped_bad_datetime += 1
                continue
            kickoff = datetime.combine(d, t)

            group_label = ""
            if mapping.get("group") is not None:
                group_label = str(r[mapping["group"]] or "").strip()
                if group_label == "-":
                    group_label = ""

            parsed.append(
                {
                    "round_number": round_number,
                    "kickoff_time": kickoff,
                    "home_team": pair[0],
                    "away_team": pair[1],
                    "group_label": group_label or None,
                }
            )

    # Fallback: если не нашли ни одной строки через заголовки —
    # пытаемся вытащить расписание эвристикой по каждой строке.
    if not parsed:
        for ws in wb.worksheets:
            sheet_rows = list(ws.iter_rows(values_only=True))
            for r_tuple in sheet_rows:
                if r_tuple is None:
                    continue
                r = list(r_tuple)
                if not any(x is not None and str(x).strip() for x in r):
                    continue

                round_idx = None
                round_number = None
                match_idx = None
                pair = None

                for idx, cell in enumerate(r):
                    if round_number is None:
                        rn = _round_to_number(cell)
                        if rn is not None:
                            round_number = rn
                            round_idx = idx
                    if pair is None:
                        p = _parse_match_pair(cell)
                        if p is not None:
                            pair = p
                            match_idx = idx

                if round_number is None:
                    continue
                if pair is None:
                    skipped_no_match += 1
                    continue

                kickoff = _extract_kickoff_from_row(r)
                if kickoff is None:
                    skipped_bad_datetime += 1
                    continue

                skip_indexes: set[int] = set()
                if round_idx is not None:
                    skip_indexes.add(round_idx)
                if match_idx is not None:
                    skip_indexes.add(match_idx)
                group_label = _detect_group_from_row(r, skip_indexes=skip_indexes)

                parsed.append(
                    {
                        "round_number": round_number,
                        "kickoff_time": kickoff,
                        "home_team": pair[0],
                        "away_team": pair[1],
                        "group_label": group_label or None,
                    }
                )

    if not parsed:
        raise RuntimeError(
            "Не смог распознать ни одной строки расписания. "
            "Проверь формат столбцов: группа | раунд | дата | время | матч."
        )

    async with SessionLocal() as session:
        tq = await session.execute(select(Tournament).where(Tournament.code == WC_CODE).limit(1))
        tournament = tq.scalar_one_or_none()
        if tournament is None:
            tournament = Tournament(
                code=WC_CODE,
                name=WC_NAME,
                round_min=1,
                round_max=9,
                is_active=1,
            )
            session.add(tournament)
            await session.flush()
        else:
            tournament.name = WC_NAME
            tournament.round_min = 1
            tournament.round_max = 9
            tournament.is_active = 1

        if clear_wc_matches:
            await session.execute(delete(Match).where(Match.tournament_id == int(tournament.id)))

        inserted = 0
        deduped = 0
        for row in parsed:
            exists_q = await session.execute(
                select(Match.id).where(
                    Match.tournament_id == int(tournament.id),
                    Match.round_number == int(row["round_number"]),
                    Match.home_team == str(row["home_team"]),
                    Match.away_team == str(row["away_team"]),
                    Match.kickoff_time == row["kickoff_time"],
                )
            )
            if exists_q.first() is not None:
                deduped += 1
                continue
            session.add(
                Match(
                    tournament_id=int(tournament.id),
                    round_number=int(row["round_number"]),
                    home_team=str(row["home_team"]),
                    away_team=str(row["away_team"]),
                    kickoff_time=row["kickoff_time"],
                    group_label=row["group_label"],
                    source="manual",
                    is_placeholder=0,
                )
            )
            inserted += 1

        await session.commit()

    print("=== WC2026 import done ===")
    print(f"file: {xlsx_path}")
    print(f"parsed_rows: {len(parsed)}")
    print(f"inserted: {inserted}")
    print(f"deduped_existing: {deduped}")
    print(f"skipped_no_match: {skipped_no_match}")
    print(f"skipped_bad_round: {skipped_bad_round}")
    print(f"skipped_bad_datetime: {skipped_bad_datetime}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import WC2026 schedule from Excel (.xlsx)")
    parser.add_argument("xlsx_path", type=str, help="Path to Excel file")
    parser.add_argument("--clear", action="store_true", help="Delete existing WC2026 matches before import")
    args = parser.parse_args()

    xlsx = Path(args.xlsx_path).expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"File not found: {xlsx}")
    asyncio.run(run_import(xlsx, clear_wc_matches=args.clear))


if __name__ == "__main__":
    main()
